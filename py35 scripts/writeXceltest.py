from openpyxl import Workbook
from openpyxl.cell import get_column_letter

wbook = Workbook()
dest_filename = "write_test.xlsx"

wsheet1 = wbook.active
wsheet1.title = "account"

wsheet2 = wbook.create_sheet(title="transactions")
wsheet2['A2'] = 'POWER'

wbook.save(dest_filename)
